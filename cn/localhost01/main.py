

# -*- coding: utf-8 -*-
import xlrd
import xlwt
from datetime import date, datetime

# from openpyxl import load_workbook

from openpyxl import Workbook

from openpyxl import *

from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import pandas as pd
import numpy as np






# def read_excel():
#
#     print("Hello World")
#     # 打开文件
#     # workbook = xlrd.open_workbook(r'D:\test5.xls')
#     wb = load_workbook(r'D:\test5.xlsx')
#     # # 读取sheetname
#     # print('输出文件所有工作表名：\n', wb.sheetnames)
#     # # 获取所有sheet
#     # print
#     # workbook.sheet_names()  # [u'sheet1', u'sheet2']
#     # sheet1_name = workbook.sheet_names()[0]
#     # sheet2_name = workbook.sheet_names()[1]
#
#     # #
#     # 根据sheet索引或者名称获取sheet内容
#     # sheet2 = workbook.sheet_by_index(0)  # sheet索引从0开始
#     # sheet2 = workbook.sheet_by_name('sheet2')
#
#     # table = wb.sheets()[0]  # 通过索引顺序获取
#     # names = workbook.sheet_names()  # 返回book中所有工作表的名字
#     # filename = r'D:\test.xls'
#     # wb = load_workbook(filename)
#     # sheet1_name = workbook.sheet_names()[0]
#     # ws = workbook.active
#     # workbook.delete_cols(3)  # 删除第 3 列数据
#     # print("delete  cols")
#     # workbook.save(r'D:\test.xls')
#     # ws = wb['test']
#     # ws2 = wb[sheet_names[0]]
#     # ws.delete_cols(1, 4)  # 从第m列开始，删除n列
#     # ws = wb.get_sheet_by_name(sheetNames[0])
#     # sheet = wb.get_sheet_by_name("第一行")
#     # sheet_names = wb.sheetnames  # 返回一个列表
#     # ws2 = wb[sheet_names[0]]  # index为0为第一张表
#     # ws2.delete_rows(1)  # 删除第n行
#     # ws2.delete_rows(2)   # 删除第n行
#     # ws2.delete_rows(3)  # 删除第n行
#     # ws2.delete_cols(1, 3)  # 从第m列开始，删除n列
#     # wb.save(r'D:\test3.xlsx')
#     # print("delete  success\n")
#
#     # data = pd.read_excel(r'D:\test3.xlsx')
#     # df = pd.DataFrame(pd.read_excel(r'D:\test5.xlsx'))
#     # print (df)
#     # sheet1 = wb.sheet_by_index(0)  # sheet索引从0开始
#     sheet_names = wb.sheetnames
#     sheet1 = wb[sheet_names[0]]  # sheet索引从0开始
#     print(sheet1.col_values(9)[5])   # 获取第四行内容,值
#     # print(sheet1.row_values(3).replace("1", '10'))
#     # print (sheet.col_values(9)[5].replace("1", '10'))
#     # wb5 = load_workbook(r'D:\test5.xlsx')
#     # print (df.sheet_names()[0])
#     # df.replace('1', 10)
#     # df.replace('regex='CSC-J' , value='test'')
#     # data[data.姓名 == '张三'].语文
#     # wb5.save(r'D:\test5.xlsx')
#     print("replace  success\n")


# if __name__ == '__main__':
#     read_excel()


# -*- coding: utf-8 -*-
import xlrd
import xlwt
from datetime import date, datetime
from xlutils.copy import copy
from xlrd import XL_CELL_NUMBER

from openpyxl.styles import numbers


def read_excel():
    filename = 'D:/未承保.xlsx'
    filename2 = 'D:/未到账.xlsx'
    wb = load_workbook(filename)
    wb2 = load_workbook(filename2)
    # Sheet1 = wb.get_sheet(0)
    # ws = wb.sheet_by_index(0)
    # print(wb.sheet_names()[0])
    ws = wb['sheet1']
    ws2 = wb2['sheet1']
    ws.title = '未承保'  # 修改名为Sheet1工作表名称
    ws2.title = '未到账'  # 修改名为Sheet1工作表名称

    wb.create_sheet(title='未到账', index=1)
    ws3 = wb2['未到账']
    # ws_sheet2 = wb2['test2']
    # print(ws_sheet2.title)
    wb.save('D:/未承保.xls')  # 保存变更
    wb2.save('D:/未到账.xls')  # 保存变更
    # wb.close()
    # wb2.close()

    # wb = xlrd.open_workbook('D:/未承保.xls')
    # wb2 = xlrd.open_workbook('D:/未到账.xls')

    # sheets1 = ws2.sheet_names()[0]    #获取sheet页未承保
    # print(sheets1)
    # sheets2 = w  File "D:/Taobao_order_robot-develop2/cn/localhost01/main.py", line 130, in read_excel
    #     max_row = sheet2.max_row  # 最大行数s2.sheet_names()[0]   # 未到账
    # sheet2 = wb.get_sheet_by_name(sheets2[0])
    # sheet1 = wb.sheet_by_index(1)
    # sheet2 = wb2.sheet_by_index(0)

    # wb = load_workbook('D:/未到账.xls')
    # wb2 = load_workbook('D:/未承保.xls')
    # ws = wb['未到账']
    # ws2 = wb2['未到账']
    #
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(ws2.iter_rows()):
        for j, cell in enumerate(row):
            ws3.cell(row=i + 1, column=j + 1, value=cell.value)

    wb.save('D:/未承保.xls')




# def combine_excel(file_path1, file_path2, target_path):
#     # 文件读入
#     data1 = pd.read_excel(file_path1)
#     data2 = pd.read_excel(file_path2)
#     target_file = pd.ExcelWriter(target_path)
#
#     data1.to_excel(test, sheet_name="未承保", index=False)
#     data2.to_excel(test, sheet_name="未到账", index=False)
#
#     # return test


# def read_excel():
#     # 打开文件
#     # workbook = xlrd.open_workbook(r'D:\test5.xlsx')
#     workbook = xlrd.open_workbook(r'D:\test5.xls')
#
#     filename = 'D:/test8.xlsx'
#     wb_test6 = load_workbook(filename)
#     new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
#     new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
#
#     # 预定一个格式
#     style = xlwt.XFStyle()
#     style.number_format = numbers.FORMAT_NUMBER
#     # 根据sheet索引或者名称获取sheet内容
#     sheet2 = workbook.sheet_by_index(0)  # sheet索引从0开始
#
#
#     # 获取单元格内容
#     sheet2.cell(12, 8).number_format = numbers.NumberFormat
#     new_write = sheet2.col_values(8)[12]
#     print(new_write)
#     new_write = int(new_write)
#     new_worksheet.write(12, 8, new_write, style)
#     # new_worksheet.write(4, 8, new_write, style)
#     print(sheet2.col_values(12, 8))
#
#     ws = wb_test6['Sheet1']
#     ws.title = '计划'  #
#     wb_test6.save(filename)  # 保存变更
#     # sheet1_name = wb_test6.sheet_names()[0]  # 获得工作表 表名
#     # print(sheet1_name)
#     # print(workbook.sheet_names()[0])
#     # ws = wb_test6[sheet1_name]
#     # ws.title = 'ni'    # 修改名为Sheet1工作表名称
#     # workbook.title = '未承保'  # 修改名为Sheet1工作表名称
#     new_workbook.save(r'D:\test5.xls')  # 保存工作簿
#     # ord(sheet2.cell_value(10, 7))
#     # new_workbook.save(r'D:\test5.xls')  # 保存工作簿
#     # print(sheet2.cell_type(7, 7))
#     # print(xlrd.XL_CELL_NUMBER('12'))
#     # print(sheet2.col_values(7)[4].ctype)
#
#     # 用for循环的方式连续向单元格中写入内容
#     # for i in range(7, 9):
#     #     for j in range(4, 359):
#     #         write = sheet2.col_values(i)[j].replace('CSC-J', '')
#     #         new_worksheet.write(j, i, write)
#     #         new_workbook.save(r'D:\test5.xls')  # 保存工作簿
#     #         print(sheet2.col_values(i)[j])
#     #         print("xls格式表格【追加】写入数据成功！")







if __name__ == '__main__':
    read_excel()



# -*- coding:utf-8 -*-
# from xlrd import open_workbook
# from xlutils.copy import copy
# import re
#
#
#
# def getrule(rfile='D:/test1.txt'):
#     try:
#         rdict = {}
#         with open(rfile, 'r') as f:
#             for line in f:
#                 rline = line.split('->')
#                 rdict[rline[0].strip()] = rline[1].strip()
#                 print rdict[rline[0].strip()]
#         return rdict
#     except Exception, e:
#         print e
#
#
# if __name__ == '__main__':
#     excelfile = 'D:/test1.xls'
#     rdict = getrule()
#     print rdict
#     rb = open_workbook(excelfile)
#     rs = rb.sheet_by_index(0)
#     wb = copy(rb)
#     ws = wb.get_sheet(0)
#     nrows = rs.nrows
#     ncols = rs.ncols
#     table = rb.sheets()[0]
#     prices = table.row_values(0)[0]
#     print prices
#     c = prices.replace('1', '88')
#     print c
#
#     strinfo = re.compile('1')
#     b = strinfo.sub('python', prices)
#     print b
#     for i in range(0,nrows):
#         for j in range(0,ncols):
#             cvalue = rs.cell(i, j).value
#             # if type(cvalue).__name__ == 'float':
#             #     cvalue = str(int(cvalue))
#             # if rdict.has_key(cvalue):
#             #     print '%s is replaced by %s' % (cvalue, rdict[cvalue])
#             #     ws.write(i, j, rdict[cvalue])
#             for repStr in rdict.keys():
#                 cvalue = cvalue.replace(repStr, rdict[repStr])
#                 ws.write(i, j, cvalue)
#     wb.save(excelfile)
#     print 'OK!'



